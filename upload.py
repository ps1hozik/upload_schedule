import os
from pandas import DataFrame
from openpyxl import load_workbook, cell as op_cell
from datetime import date, datetime, timedelta

from config import get_database

MONTHS = {
    "янв": 1,
    "фев": 2,
    "мар": 3,
    "апр": 4,
    "мая": 5,
    "май": 5,
    "июн": 6,
    "июл": 7,
    "авг": 8,
    "сен": 9,
    "окт": 10,
    "ноя": 11,
    "дек": 12,
}
DAYS = ["пон", "вто", "сре", "чет", "пят", "суб"]


def date_format(date_string: str):
    try:
        if date_string != "" and date_string != "None":
            date_parts = date_string.split()
            day = int(date_parts[0])
            month = MONTHS[date_parts[1][:3].lower()]
            year = int(date_parts[2])
            return str(date(year, month, day))
        return date_string
    except Exception as exc:
        error = f"\n[?]date_format: {exc}"
        print(error)
        return


def get_data(filename: str):
    try:
        wb = load_workbook(filename=filename)
        ws = wb.active
        data = []
        empty_index = []
        next_i = 0
        for i, row in enumerate(ws.iter_rows(), start=1):
            if next_i == 0:
                for cell in row:
                    if "курс" in str(cell.value).strip():
                        next_i = i + 1
                        break
                else:
                    continue
            if next_i == i:
                continue
            row_data = []
            for cell in row:
                if isinstance(cell, op_cell.cell.MergedCell):
                    for r in ws.merged_cells.ranges:
                        if cell.coordinate in r:
                            row_data.append(str(r.start_cell.value).strip())
                            break
                else:
                    cell_value = (
                        str(cell.value).strip() if cell.value is not None else ""
                    )
                    row_data.append(cell_value)
            if len(data) > 2:
                if row_data.count("") != len(row_data) and (
                        row_data[0] == "" or row_data[1] == "" or row_data[2] == ""
                ):
                    empty_index.append(i - 12)
            if row_data[0] != "" and (
                    row_data.count(row_data[0]) == len(row_data)
                    or row_data.count(row_data[0]) == len(row_data) - 1
            ):
                row_data = [""] * len(row_data)
            data.append(row_data)

        for i in empty_index:
            d1, d2, d3 = data[i][:3]
            if d1 == "":
                for d in data[3:]:
                    if (d2 or d3) in d and d1 != d[0]:
                        data[i][0] = d[0]
                        break
            if d2 == "":
                for d in data[3:]:
                    if (d1 or d3) in d and d2 != d[1]:
                        data[i][1] = d[1]
                        break
            if d3 == "":
                for d in data[3:]:
                    if (d1 or d2) in d and d3 != d[2]:
                        data[i][2] = d[2]
                        break
        index = 0
        for i, d in enumerate(data):
            if len(d[0]) > 2 and str(d[0]).lower()[:3] in DAYS:
                index = i
                break

        courses = [i[0] if i is not None and len(i) > 1 else i for i in data[0]]
        groups = data[1]
        subgroups = data[2]
        schedule = data[index:-3]
        subgroups[0], subgroups[1], subgroups[2] = 0, 1, 2

        result_dict = {}
        while courses[-1] == "":
            del groups[-1]
            del subgroups[-1]
            del courses[-1]
            for s in schedule:
                del s[-1]
        df = DataFrame(schedule, columns=subgroups)
        for i in range(3, len(groups)):
            course = courses[i]
            group = groups[i]
            subgroup = subgroups[i]

            if group in result_dict:
                result_dict[group]["sub_groups"].append(subgroup)
            else:
                result_dict[group] = {
                    "group_name": group,
                    "course": course,
                    "sub_groups": [subgroup],
                }
        return df, list(result_dict.values())
    except Exception as exc:
        error = f"\n[?]get_dataFrame: {exc}"
        print(error)
        return None, None


def get_schedule(
        df: DataFrame, groups_list: list, faculty_name: str, errors: list
) -> list | None:
    schedule_dicts_list = []
    try:
        group: dict
        for group in groups_list:
            i = 0
            for subgroup in group.get("sub_groups"):
                lessons = []
                schedule_list = []
                schedule_dict = {"group_name": subgroup}

                while df[2][i] != "":
                    lesson = list(df[subgroup][i: i + 3])
                    if lesson[0] != "" or lesson[1] != "" or lesson[2] != "":
                        if lesson[0] == lesson[1]:
                            lesson[1] = ""
                        lesson_with_times = {
                            "lesson": {
                                "name": lesson[0],
                                "teacher": lesson[1],
                                "auditorium": lesson[2],
                            },
                            "time": df[2][i + 1],
                            "number": df[2][i],
                        }
                    else:
                        lesson_with_times = {
                            "lesson": None,
                            "time": df[2][i + 1],
                            "number": df[2][i],
                        }
                    lessons.append(lesson_with_times)
                    i += 3
                    if i >= df.shape[0] or df[2][i] == "":
                        for lesson in lessons:
                            if lesson["lesson"] is not None:
                                one_day_lessons_dict = {
                                    "day": df[0][i - 2],
                                    "date": date_format(df[1][i - 2]),
                                    "lessons": lessons,
                                }
                                schedule_list.append(one_day_lessons_dict)
                                break
                        i += 1
                        lessons = []
                        if i >= df.shape[0]:
                            i -= 1
                            break

                schedule_dict["schedule"] = schedule_list
                schedule_dicts_list.append(schedule_dict)
                i = 0
        return schedule_dicts_list
    except Exception as exc:
        error = (
            f"\n[?]get_schedule: {faculty_name}\n{schedule_dicts_list[-1]=} \n {exc}"
        )
        errors.append(error)
        print(error)
        return schedule_dicts_list


def get_saturday():
    now = datetime.now()
    weekday = now.weekday()
    if weekday != 4 and weekday != 5:
        return None
    sat = now + timedelta(days=5 - weekday)
    return str(sat.date())


def upload_data(
        faculty_name: str, data: list, type: str, sat: str, file: str, errors: list
) -> None:
    collection = f"{type} {faculty_name}"
    try:
        dbname = get_database()
        collection_data = dbname[collection]
        if sat:
            if "Расписание" in type:
                for d in data:
                    lessons = collection_data.find_one({"group_name": d["group_name"]})
                    if lessons and lessons.get("schedule"):
                        for l in lessons["schedule"]:
                            if l["date"] == sat:
                                d["schedule"].insert(0, l)
                                break

        collection_data.drop()
        collection_data = dbname[collection]

        collection_data.insert_many(data)

        print(f"[+]Successfully uploaded: {collection}")
    except Exception as exc:
        error = f"\n[?]upload_data222: {file}\n{exc}"
        errors.append(error)
        print(error)


def upload():
    data_folder = os.path.join("data")
    sat = get_saturday()
    errors = []
    for root, _, files in os.walk(data_folder):
        groups_list = []
        schedule_list = []
        faculty_name = ""
        for f in files:
            file = os.path.join(root, f)
            faculty_name = file.split(os.sep)[1]
            schedule_df, groups = get_data(file)
            if schedule_df is None or groups is None:
                continue
            groups_list.extend(groups)
            sch = get_schedule(schedule_df, groups, faculty_name, errors=errors)
            if sch is None:
                continue
            schedule_list.extend(sch)
        if groups_list and schedule_list:
            upload_data(
                faculty_name,
                groups_list,
                "Группы",
                sat,
                file,
                errors=errors,
            )
            upload_data(
                faculty_name,
                schedule_list,
                "Расписание",
                sat,
                file,
                errors=errors,
            )
    return errors


if __name__ == "__main__":
    upload()
